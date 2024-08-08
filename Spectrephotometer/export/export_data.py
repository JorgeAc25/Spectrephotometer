import openpyxl

def export_data(name, values=[]):
    nombre_archivo = "Mediciones.xlsx"

    try:
        libro_excel = openpyxl.load_workbook(nombre_archivo)
    except FileNotFoundError:
        libro_excel = openpyxl.Workbook()

    hoja = libro_excel.active


    if not hoja['A1'].value:  
        encabezados = ["SOLUCIÓN", "400nm", "460nm", "515nm", "603nm", "625nm"]
        for col_num, encabezado in enumerate(encabezados, 1):
            hoja.cell(row=1, column=col_num, value=encabezado)

  
    ultima_fila = hoja.max_row + 1

 
    hoja.cell(row=ultima_fila, column=1, value=name)
    for col_num, dato in enumerate(values, 2):  
        hoja.cell(row=ultima_fila, column=col_num, value=dato)

    libro_excel.save(nombre_archivo)
    print(f"Datos para {name} agregados con éxito")

    libro_excel.close()
